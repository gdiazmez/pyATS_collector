from genie.testbed import load
import argparse
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import logging
import os
import re
import getpass
from utils import Create_Testbed,pool_connection
from rich.progress import Progress

logging.basicConfig(filename='log_full_info.log', level=logging.DEBUG,
                    format='%(asctime)s %(levelname)s : %(message)s')


def Testbed_routine(hostname):
    progress.update(task1, advance=1)
    uut = tb.devices[hostname]
    connected=False
    if jump_replace:
        try:
            jump = tb.devices['jump_host']
            jump.connect(log_stdout=verbose_flag)
            uut_ip = str(uut.connections['cli']['ip'])
            command = 'ssh-keygen -R {}'.format(uut_ip)
            jump.execute(command)
            print("IP {} SSH keys cleared, continue with connection".format(uut_ip))
            logging.debug("IP {} SSH keys cleared, continue with connection".format(uut_ip))
        except Exception as e:
            print("Exception catched on SSH keys clearing on jumphost")
            logging.debug("Exception catched on SSH keys clearing on jumphost:\n{}".format(e))
    try:
        uut.connect(init_config_commands=[],connection_timeout=5,log_stdout=verbose_flag)
        connected=True
    except:
        logging.debug("Error #1 in the connection to: '{}'".format(hostname))

    if not connected:
        try:
            uut.connect(init_config_commands=[],connection_timeout=5,log_stdout=verbose_flag)
            connected=True
        except:
            logging.debug("Error #2 in the connection to: '{}'".format(hostname))

    if not connected:
        try:
            for pattern in vmc_patterns:
                if pattern in hostname:
                    #Last resort creds from pattern
                    uut.credentials['default'] = dict(username='dummy', password='change_me')
            else:
                # Other group last resort creds
                uut.credentials['default'] = dict(username='etc', password='cool_pass')
            uut.connect(init_config_commands=[],connection_timeout=5,log_stdout=verbose_flag)
            connected=True
        except:
            logging.debug("Error #3 in the connection to: '{}'".format(hostname))
            progress.update(task3, advance=1)
            print("Error #3 in the connection to: '{}'. Check details".format(hostname))
            failed.append(hostname)

    if connected:
        print('Device: "{}" connected, collecting'.format(hostname))
        try:
            ### Get IPv4 from Lo0
            output = uut.parse('show ipv4 interface brief | i Loopback0')
            lo0_ipv4 = output['interface']['Loopback0']['ip_address']
            logging.info("Lo0 IPv4 on vRouter: {} is: {}".format(hostname,lo0_ipv4))
            ### Get IPv6 from Lo0
            lo0_ipv6 = ''
            try:
                output = uut.parse('show ipv6 interface Loopback0')
                for key in output['Loopback0']['ipv6'].keys():
                    if '/128' in key:
                        lo0_ipv6 = output['Loopback0']['ipv6'][key]['ipv6']
                        break
            except Exception as e:
                logging.info("Exception catched at Lo0 IPv6 \n {}".format(e))
                lo0_ipv6 = 'Not Configured'
            logging.info("Lo0 IPv6 on vRouter: {} is: {}".format(hostname,lo0_ipv6))
            ### Get Prefix-SID from Lo0
            output = uut.execute('show isis segment-routing label table | i Loopback0')
            label_re = re.compile(r'^([0-9]+\S)', re.M)
            labels = label_re.findall(output)
            prefix_sid = ''
            if labels:
                prefix_sid = labels[0]
            else:
                prefix_sid = 'No Prefix-SID'
            logging.info("Prefix-SID vRouter: {} is: {}".format(hostname,prefix_sid))
            ### Get IPv4 from Lo10
            try:
                output = uut.parse('show ipv4 interface brief | i Loopback10')
                lo10_ipv4 = output['interface']['Loopback10']['ip_address']
            except Exception as e:
                logging.info("Exception catched at Lo10 ipv4")
                lo10_ipv4 = 'Not present'
            logging.info("Lo10 IPv4 on vRouter: {} is: {}".format(hostname,lo10_ipv4))
            ### Get IPv6 from Lo10
            lo10_ipv6 = ''
            try:
                vrf = 'NETWORK-MGMT'
                sddc_patterns = ['RANMK', 'RANMN']
                for pattern in sddc_patterns:
                    if pattern in hostname:
                        vrf = 'SERVICE-MGMT'
                        break
                output = uut.parse('show ipv6 vrf {} interface Loopback10'.format(vrf))
                for key in output['Loopback10']['ipv6'].keys():
                    if '/128' in key:
                        lo10_ipv6 = output['Loopback10']['ipv6'][key]['ipv6']
                        break
            except Exception as e:
                logging.info("Exception catched at Lo10 IPv6 \n {}".format(e))
                lo10_ipv6 = 'Not Configured'
            logging.info("Lo10 IPv6 on vRouter: {} is: {}".format(hostname,lo10_ipv6))
            ### Get XR release
            output = uut.parse('show version')
            version = output['software_version']
            logging.info("Version on vRouter: {} is: {}".format(hostname,version))

            ### Get SMU in Commited Package List
            output = uut.execute('show install commit summary')
            logging.info("Full packages on hostname: {} is:\n{}".format(hostname,output))
            smu_state = True
            pending = 'SMU pending: \r'


            for pattern in vmc_patterns:
                if pattern in hostname:
                    smu_list = smu_list_vmc
                    break
            else:
                smu_list = smu_list_aws

            for smu in smu_list:
                if smu in output:
                    continue
                else:
                    pending = pending + smu + " \r"
                    smu_state = False
            if smu_state:
                smu_state="OK"
            else:
                smu_state=pending

            ### Get License Registration Status
            lic_status=""
            output = uut.execute('show license status | i Status:',timeout=10)
            if "Status" in output:
                if "UNREGISTERED" in output:
                    lic_status="Unregistered"
                else:
                    lic_status="Registered"
            else:
                lic_status="No Status Info. Check vRouter"
            # Sheet Format: "Hostname","Lo0 - IPv4","Lo0 - IPv6","Lo0 - Prefix SID","Lo10 - IPv4","Lo10 - IPv4","XR Version","SMU State","License Status"
            sheet.append([hostname,lo0_ipv4,lo0_ipv6,prefix_sid,lo10_ipv4,lo10_ipv6,version,smu_state,lic_status])
        except Exception as e:
            logging.debug('Exception catched on Hostname: "{}"\n{}'.format(hostname,e))
        uut.destroy()
        progress.update(task2, advance=1)
        logging.debug('Processing complete for device: "{}"'.format(hostname))
        print('Processing complete for device: "{}"'.format(hostname))

def main():
    parser=argparse.ArgumentParser()
    parser.add_argument('-f','--file', help='Input File', required=True)
    parser.add_argument('-c','--credentials', help='Send vRouter creds in format user:pass', required=False)
    parser.add_argument('-j','--jump', action='store_true', help='Flag to ask for SSH Jumphost detail', required=False)
    parser.add_argument('-d','--detail', action='store_true', help='Flag to print verbose output from pyATS', required=False)
    parser.add_argument('-e','--environment', action='store_true', help='Flag to use env_variables', required=False)
    parser.add_argument('-r','--replace', action='store_true', help='Flag to use ssh-keygen -R to update ssh keys if jumphost is used', required=False)
    args=parser.parse_args()

    file = args.file
    jump_bool = False
    global jump_replace
    jump_replace = False
    jump_ip = ""
    jump_port = ""
    jump_user = ""
    jump_pass = ""
    if args.jump :
        jump_bool = True
        if not args.environment:
            jump_ip = input("Enter Jumphost IP: ")
            jump_port = input("Enter Jumphost Port: ")
            jump_user = input("Enter Jumphost Username: ")
            jump_pass = getpass.getpass('Enter Jumphost Password:')
        else:
            message='''
Using environment variables for Jumphost. Check these if connection fails
    JUMP_IP
    JUMP_PORT
    JUMP_USER
    JUMP_PASS'''
            print(message)
            jump_ip = os.environ.get('JUMP_IP')
            jump_port = os.environ.get('JUMP_PORT')
            jump_user = os.environ.get('JUMP_USER')
            jump_pass = os.environ.get('JUMP_PASS')
        if args.replace:
            jump_replace = True

    if args.environment:
        user = os.environ.get('VROUTER_USER')
        password = os.environ.get('VROUTER_PASS')
        message='''
Using environment variables for device credentials. Check these if connection fails
    VROUTER_USER
    VROUTER_PASS
        '''
        print(message)
    elif args.credentials:
        creds = args.credentials
        user = creds.split(":")[0]
        password = creds.split(":")[1]
    else:
        print("Credential argument or environment not found, enter vRouter credentials/n")
        user = input("Enter vRouter Username:")
        password = getpass.getpass('Enter vRouter Password:')

    global full_table
    global smu_list
    global smu_list_vmc
    global smu_list_aws
    smu_list_aws=[
        # Smart License SMU
        'CSCwa80752',
        # ISIS SMU
        'CSCwd63200',
        # Memory Leak SMU
        'CSCwb74098',
        # Throttling SMU
        'CSCwf27917'
    ]
    smu_list_vmc=[
            # Memory Leak SMU
            'CSCwb74098',
            # ISIS SMU
            'CSCwd63200',
            # Throttling SMU
            'CSCwf27917'
        ]
    smu_list = smu_list_aws
    global vmc_patterns
    vmc_patterns = ['VMC0','VNF0','RANMK', 'RANMN']
    global verbose_flag
    if args.detail:
        verbose_flag=True
    else:
        verbose_flag=False
    full_table = load_workbook(file)
    devices = full_table['Device Info']

    device_ips = []
    hostnames = []

    for i in range(1, len(list(list(devices.columns)[0]))):
        hostnames.append(list(list(devices.columns)[0])[i].value)

    for i in range(1, len(list(list(devices.columns)[1]))):
        device_ips.append(list(list(devices.columns)[1])[i].value)

    testbed = Create_Testbed(user,password,hostnames,device_ips,jump_bool,
                             jump_ip,jump_port,jump_user,jump_pass)
    try:
        global tb
        tb = load(testbed)
    except Exception as e:
        logging.debug('Exception on load testbed:\n {}'.format(e))

    os.system('rm Output.xlsx')
    global sheet
    out_wb = Workbook()
    sheet = out_wb.active
    sheet.title = "Parsed"
    ### 1st row
    sheet.append(["Hostname","Lo0 - IPv4","Lo0 - IPv6","Lo0 - Prefix SID","Lo10 - IPv4","Lo10 - IPv4","XR Version","SMU State","License Status"])

    global total_devices
    global progress
    global task1
    global task2
    global task3
    global failed
    total_devices = len(hostnames)
    failed = []
    print('Going to process: {} devices'.format(total_devices))
    with Progress() as progress:
        task1 = progress.add_task("[cyan]Processed...", total=total_devices)
        task2 = progress.add_task("[green]Collected...", total=total_devices)
        task3 = progress.add_task("[red]Failed...", total=total_devices)
        result = pool_connection(12,hostnames,Testbed_routine)
    if result:
        if len(failed)>0:
            print("\nRoutine ended with failed devices:")
            for host in failed:
                print("{}".format(host))
        else:
            print("Success!!!")
        out_wb.save(filename = "Output.xlsx")
        print("\nResults on ./Outputx.xlsx")

if __name__ == "__main__":
    main()

