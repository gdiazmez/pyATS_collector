from genie.testbed import load
import argparse
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import logging
import os
import re
import getpass
from utils import Create_Gen_Testbed,pool_connection
from rich.progress import Progress

logging.basicConfig(filename='snmp_info.log', level=logging.DEBUG,
                    format='%(asctime)s %(levelname)s : %(message)s')


def Testbed_routine(hostname):
    progress.update(task1, advance=1)
    uut = tb.devices[hostname]
    reachable=False
    connected=False
    if jump_replace:
        try:
            jump = tb.devices['jump_host']
            jump.connect(log_stdout=verbose_flag)
            uut_ip = str(uut.connections['cli']['ip'])
            command = 'ssh-keygen -R {}'.format(uut_ip)
            jump.execute(command)
            print("IP {} SSH keys cleared, continue with connection".format(uut_ip))
            logging.debug("IP {} SSH keys cleared, continue with connection".format(uut_ip))
        except Exception as e:
            print("Exception catched on SSH keys clearing on jumphost")
            logging.debug("Exception catched on SSH keys clearing on jumphost:\n{}".format(e))
    ### Trying Ping from CSPC
    jump = tb.devices['jump_host']
    jump.connect(log_stdout=verbose_flag)
    uut_ip = str(uut.connections['cli']['ip'])
    command = 'ping -W 1 -c 1 {}'.format(uut_ip)
    output = jump.execute(command)
    if '1 received' in output:
        print("Device {} reachable, flagging to connect using generic OS and learn".format(hostname))
        reachable = True
    if reachable:
        try:
            uut.connect(init_config_commands=[],connection_timeout=5,
                        log_stdout=verbose_flag,learn_os=True,learn_hostname=True)
            connected=True
        except:
            logging.debug("Error #1 in the connection to: '{}'".format(hostname))

        if not connected:
            try:
                uut.connect(init_config_commands=[],connection_timeout=5,
                        log_stdout=verbose_flag,learn_os=True,learn_hostname=True)
                connected=True
            except:
                logging.debug("Error #2 in the connection to: '{}'".format(hostname))

        if not connected:
            try:
                uut.credentials['default'] = dict(username='sdnc.aws', password='D1$HmgmtAW$')
                uut.connect(init_config_commands=[],connection_timeout=5,log_stdout=verbose_flag)
                connected=True
            except:
                logging.debug("Error #3 in the connection to: '{}'".format(hostname))
                progress.update(task3, advance=1)
                print("Error #3 in the connection to: '{}'. Check details".format(hostname))
                failed.append(hostname)

        if connected:
            print('Device: "{}" connected, working using OS: {}'.format(hostname,uut.os))
            ### Develop code using uut.os to segregate IOS-XE and IOS-XR commands
            uut.destroy()
            progress.update(task2, advance=1)
            logging.debug('Processing complete for device: "{}"'.format(hostname))
            print('Processing complete for device: "{}"'.format(hostname))
    else:
        logging.debug('Device: "{}" not reachable, skiping it'.format(hostname))
        progress.update(task3, advance=1)
        print('Device: "{}" not reachable, skiping it'.format(hostname))
        failed.append(hostname)

def main():
    parser=argparse.ArgumentParser()
    parser.add_argument('-f','--file', help='Input File', required=True)
    parser.add_argument('-c','--credentials', help='Send vRouter creds in format user:pass', required=False)
    parser.add_argument('-j','--jump', action='store_true', help='Flag to ask for SSH Jumphost detail', required=False)
    parser.add_argument('-d','--detail', action='store_true', help='Flag to print verbose output from pyATS', required=False)
    parser.add_argument('-e','--environment', action='store_true', help='Flag to use env_variables', required=False)
    parser.add_argument('-r','--replace', action='store_true', help='Flag to use ssh-keygen -R to update ssh keys if jumphost is used', required=False)
    args=parser.parse_args()

    file = args.file
    jump_bool = False
    global jump_replace
    jump_replace = False
    jump_ip = ""
    jump_port = ""
    jump_user = ""
    jump_pass = ""
    if args.jump :
        jump_bool = True
        if not args.environment:
            jump_ip = input("Enter Jumphost IP: ")
            jump_port = input("Enter Jumphost Port: ")
            jump_user = input("Enter Jumphost Username: ")
            jump_pass = getpass.getpass('Enter Jumphost Password:')
        else:
            message='''
Using environment variables for Jumphost. Check these if connection fails
    JUMP_IP
    JUMP_PORT
    JUMP_USER
    JUMP_PASS'''
            print(message)
            jump_ip = os.environ.get('JUMP_IP')
            jump_port = os.environ.get('JUMP_PORT')
            jump_user = os.environ.get('JUMP_USER')
            jump_pass = os.environ.get('JUMP_PASS')
        if args.replace:
            jump_replace = True

    if args.environment:
        user = os.environ.get('VROUTER_USER')
        password = os.environ.get('VROUTER_PASS')
        message='''
Using environment variables for device credentials. Check these if connection fails
    VROUTER_USER
    VROUTER_PASS
        '''
        print(message)
    elif args.credentials:
        creds = args.credentials
        user = creds.split(":")[0]
        password = creds.split(":")[1]
    else:
        print("Credential argument or environment not found, enter vRouter credentials/n")
        user = input("Enter vRouter Username:")
        password = getpass.getpass('Enter vRouter Password:')

    global full_table
    global verbose_flag
    if args.detail:
        verbose_flag=True
    else:
        verbose_flag=False
    full_table = load_workbook(file)
    devices = full_table['Device Info']

    device_ips = []
    hostnames = []

    for i in range(1, len(list(list(devices.columns)[0]))):
        hostnames.append(list(list(devices.columns)[0])[i].value)

    for i in range(1, len(list(list(devices.columns)[1]))):
        device_ips.append(list(list(devices.columns)[1])[i].value)

    testbed = Create_Gen_Testbed(user,password,hostnames,device_ips,jump_bool,
                             jump_ip,jump_port,jump_user,jump_pass)
    try:
        global tb
        tb = load(testbed)
    except Exception as e:
        logging.debug('Exception on load testbed:\n {}'.format(e))

    os.system('rm Output.xlsx')
    global sheet
    out_wb = Workbook()
    sheet = out_wb.active
    sheet.title = "Parsed"
    ### 1st row
    sheet.append(["Hostname","Lo0 - IPv4","Lo0 - IPv6","Lo0 - Prefix SID","Lo10 - IPv4","Lo10 - IPv4","XR Version","SMU State","License Status"])

    global total_devices
    global progress
    global task1
    global task2
    global task3
    global failed
    total_devices = len(hostnames)
    failed = []
    print('Going to process: {} devices'.format(total_devices))
    with Progress() as progress:
        task1 = progress.add_task("[cyan]Processed...", total=total_devices)
        task2 = progress.add_task("[green]Collected...", total=total_devices)
        task3 = progress.add_task("[red]Failed...", total=total_devices)
        result = pool_connection(12,hostnames,Testbed_routine)
    if result:
        if len(failed)>0:
            print("\nRoutine ended with failed devices:")
            for host in failed:
                print("{}".format(host))
        else:
            print("Success!!!")
        out_wb.save(filename = "Output.xlsx")
        print("\nResults on ./Outputx.xlsx")

if __name__ == "__main__":
    main()

