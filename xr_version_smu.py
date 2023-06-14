from genie.testbed import load
import argparse
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import logging
import os
import getpass
from utils import Create_Testbed,pool_connection

logging.basicConfig(filename='log_collect.log', level=logging.DEBUG,
                    format='%(asctime)s %(levelname)s : %(message)s')


def Testbed_routine(hostname):

    uut = tb.devices[hostname]
    connected=False
    if jump_replace:
        try:
            jump = tb.devices['jump_host']
            jump.connect(log_stdout=verbose_flag)
            uut_ip = str(uut.connections['cli']['ip'])
            command = 'ssh-keygen -R {}'.format(uut_ip)
            jump.execute(command)
            print("IP {} SSH keys cleared, continue with connection".format(uut_ip))
            logging.debug("IP {} SSH keys cleared, continue with connection".format(uut_ip))
        except Exception as e:
            print("Exception catched on SSH keys clearing on jumphost")
            logging.debug("Exception catched on SSH keys clearing on jumphost:\n{}".format(e))
    try:
        uut.connect(init_config_commands=[],connection_timeout=5,log_stdout=verbose_flag)
        connected=True
    except:
        logging.debug("Error #1 in the connection to: '{}'".format(hostname))

    if not connected:
        try:
            uut.connect(init_config_commands=[],connection_timeout=5,log_stdout=verbose_flag)
            connected=True
        except:
            logging.debug("Error #2 in the connection to: '{}'".format(hostname))

    if not connected:
        try:
            if vmc_flag:
                uut.credentials['default'] = dict(username='root', password='dishcisco')
            else:
                uut.credentials['default'] = dict(username='sdnc.aws', password='D1$HmgmtAW$')
            uut.connect(init_config_commands=[],connection_timeout=5,log_stdout=verbose_flag)
            connected=True
        except:
            logging.debug("Error #3 in the connection to: '{}'".format(hostname))
            print("Error #3 in the connection to: '{}'. Check details".format(hostname))

    if connected:
        print('Device: "{}" connected, collecting'.format(hostname))
        try:
            ### Get XR release
            output = uut.parse('show version')
            version = output['software_version']
            logging.info("Version on vRouter: {} is: {}".format(hostname,version))

            ### Get SMU in Commited Package List
            output = uut.execute('show install commit summary')
            logging.info("Full packages on hostname: {} is:\n{}".format(hostname,output))
            smu_state = True
            pending = 'SMU pending: \r'

            for smu in smu_list:
                if smu in output:
                    continue
                else:
                    pending = pending + smu + " \r"
                    smu_state = False
            if smu_state:
                smu_state="OK"
            else:
                smu_state=pending

            ### Get License Registration Status
            lic_status=""
            output = uut.execute('show license status | i Status:',timeout=10)
            if "Status" in output:
                if "UNREGISTERED" in output:
                    lic_status="Unregistered"
                else:
                    lic_status="Registered"
            else:
                lic_status="No Status Info. Check vRouter"
            sheet.append([hostname,version,smu_state,lic_status])
        except Exception as e:
            logging.debug('Exception catched on Hostname: "{}"\n{}'.format(hostname,e))
        print('Processing complete for device: "{}"'.format(hostname))

def main():
    parser=argparse.ArgumentParser()
    parser.add_argument('-f','--file', help='Input File', required=True)
    parser.add_argument('-c','--credentials', help='Send vRouter creds in format user:pass', required=False)
    parser.add_argument('-j','--jump', action='store_true', help='Flag to ask for SSH Jumphost detail', required=False)
    parser.add_argument('-v','--vmc', action='store_true', help='Flag to ommit check CSCwa80752 on VMC vRouters', required=False)
    parser.add_argument('-d','--detail', action='store_true', help='Flag to print verbose output from pyATS', required=False)
    parser.add_argument('-e','--environment', action='store_true', help='Flag to use env_variables', required=False)
    parser.add_argument('-r','--replace', action='store_true', help='Flag to use ssh-keygen -R to update ssh keys if jumphost is used', required=False)
    args=parser.parse_args()

    file = args.file
    jump_bool = False
    global jump_replace
    jump_replace = False
    jump_ip = ""
    jump_port = ""
    jump_user = ""
    jump_pass = ""
    if args.jump :
        jump_bool = True
        if not args.environment:
            jump_ip = input("Enter Jumphost IP: ")
            jump_port = input("Enter Jumphost Port: ")
            jump_user = input("Enter Jumphost Username: ")
            jump_pass = getpass.getpass('Enter Jumphost Password:')
        else:
            message='''
Using environment variables for Jumphost. Check these if connection fails
    JUMP_IP
    JUMP_PORT
    JUMP_USER
    JUMP_PASS'''
            print(message)
            jump_ip = os.environ.get('JUMP_IP')
            jump_port = os.environ.get('JUMP_PORT')
            jump_user = os.environ.get('JUMP_USER')
            jump_pass = os.environ.get('JUMP_PASS')
        if args.replace:
            jump_replace = True

    if args.environment:
        user = os.environ.get('VROUTER_USER')
        password = os.environ.get('VROUTER_PASS')
        message='''
Using environment variables for device credentials. Check these if connection fails
    VROUTER_USER
    VROUTER_PASS
        '''
        print(message)
    elif args.credentials:
        creds = args.credentials
        user = creds.split(":")[0]
        password = creds.split(":")[1]
    else:
        print("Credential argument or environment not found, enter vRouter credentials/n")
        user = input("Enter vRouter Username:")
        password = getpass.getpass('Enter vRouter Password:')

    global full_table
    global smu_list
    smu_list=[
        # Smart License SMU
        'CSCwa80752',
        # ISIS SMU
        'CSCwd63200',
        # Memory Leak SMU
        'CSCwb74098',
        # Throttling SMU
        'CSCwf27917'
    ]
    global vmc_flag
    vmc_flag = False
    if args.vmc:
        vmc_flag = True
        print("VMC Flag used, Skip check on AWS license SMU\n")
        smu_list=[
            # Memory Leak SMU
            'CSCwb74098',
            # ISIS SMU
            'CSCwd63200',
            # Throttling SMU
            'CSCwf27917'
        ]
    global verbose_flag
    if args.detail:
        verbose_flag=True
    else:
        verbose_flag=False
    full_table = load_workbook(file)
    devices = full_table['Device Info']

    device_ips = []
    hostnames = []

    for i in range(1, len(list(list(devices.columns)[0]))):
        hostnames.append(list(list(devices.columns)[0])[i].value)

    for i in range(1, len(list(list(devices.columns)[1]))):
        device_ips.append(list(list(devices.columns)[1])[i].value)

    testbed = Create_Testbed(user,password,hostnames,device_ips,jump_bool,
                             jump_ip,jump_port,jump_user,jump_pass)
    try:
        global tb
        tb = load(testbed)
    except Exception as e:
        logging.debug('Exception on load testbed:\n {}'.format(e))

    os.system('rm Output.xlsx')
    global sheet
    out_wb = Workbook()
    sheet = out_wb.active
    sheet.title = "Parsed"
    ### 1st row
    sheet.append(["Hostname","XR Version","SMU State","License Status"])

    result = pool_connection(12,hostnames,Testbed_routine)
    if result:
        print("Success")
        out_wb.save(filename = "Output.xlsx")

if __name__ == "__main__":
    main()

